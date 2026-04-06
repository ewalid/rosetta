"""FastAPI application for Rosetta translation service."""

import os
import tempfile
import base64
import queue
import threading
from pathlib import Path
from typing import Optional

from dotenv import load_dotenv
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from starlette.middleware.base import BaseHTTPMiddleware
from sse_starlette.sse import EventSourceResponse
from openpyxl import load_workbook
import json
import asyncio

from rosetta.services.translation_service import count_cells, translate_file
from rosetta.services import ExcelExtractor
from .mcp import router as mcp_router, COST_PER_1000_CELLS_USD

# Load environment variables from .env file
load_dotenv()

# CORS configuration - allow frontend origins
FRONTEND_URL = os.getenv("FRONTEND_URL", "")
CORS_ALLOW_ALL = os.getenv("CORS_ALLOW_ALL", "false").lower() == "true"

ALLOWED_ORIGINS = [
    "http://localhost:3000",
    "http://127.0.0.1:3000",
]
if FRONTEND_URL:
    ALLOWED_ORIGINS.append(FRONTEND_URL)

# For corporate firewalls, allow all origins if configured
if CORS_ALLOW_ALL:
    ALLOWED_ORIGINS = ["*"]

# Limits
# Keep in sync with frontend validation/copy (50MB).
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB
MAX_CELLS = 5000

app = FastAPI(
    title="Rosetta",
    description="Excel translation API that preserves formatting, formulas, and data integrity",
    version="0.1.0",
)


class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    """Middleware to add security headers for corporate firewall compatibility."""

    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        
        # Security headers that corporate firewalls expect
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["X-XSS-Protection"] = "1; mode=block"
        response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
        response.headers["Permissions-Policy"] = "geolocation=(), microphone=(), camera=()"
        
        # HSTS header (only add if HTTPS)
        if request.url.scheme == "https":
            response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
        
        # Connection keep-alive for corporate proxies
        response.headers["Connection"] = "keep-alive"
        
        return response


# Add security headers middleware first (before CORS)
app.add_middleware(SecurityHeadersMiddleware)

# CORS middleware for frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=not CORS_ALLOW_ALL,  # Disable credentials if allowing all origins
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],
)

# Register MCP router
app.include_router(mcp_router)


@app.get("/")
async def root() -> dict:
    """Health check endpoint."""
    return {"status": "ok", "service": "rosetta"}


@app.get("/health")
async def health() -> dict:
    """Lightweight health check endpoint for firewalls and monitoring.
    
    Returns minimal JSON response quickly to help corporate firewalls
    validate the connection without heavy processing.
    """
    return {"status": "healthy"}


@app.post("/estimate")
async def estimate_cost(
    file: UploadFile = File(..., description="Excel file to estimate translation cost for"),
    sheets: Optional[str] = Form(None, description="Comma-separated sheet names (all if omitted)"),
) -> dict:
    """Estimate the cost and time for translating an Excel file.
    
    Returns cell count, estimated API cost, and estimated processing time.
    """
    # Validate file type
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")

    if not file.filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(
            status_code=400,
            detail="Invalid file type. Only Excel files (.xlsx, .xlsm, .xltx, .xltm) are supported",
        )

    # Read file content
    content = await file.read()

    # Check file size
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=400,
            detail=f"File too large. Maximum size is {MAX_FILE_SIZE // (1024 * 1024)}MB",
        )

    # Parse sheets parameter
    sheets_set = None
    if sheets:
        sheets_set = {s.strip() for s in sheets.split(",") if s.strip()}

    # Save to temp file for processing
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_input:
        tmp_input.write(content)
        input_path = Path(tmp_input.name)

    try:
        # Count cells
        cell_count = count_cells(input_path, sheets_set)
        
        if cell_count == 0:
            raise HTTPException(
                status_code=400,
                detail="No translatable content found in the file",
            )

        # Estimate cost
        estimated_cost = (cell_count / 1000) * COST_PER_1000_CELLS_USD

        # Estimate time (roughly 50 cells per second with batching)
        estimated_seconds = cell_count / 50
        if estimated_seconds < 60:
            time_estimate_seconds = int(estimated_seconds)
            time_estimate_display = f"{time_estimate_seconds} seconds"
        else:
            time_estimate_minutes = estimated_seconds / 60
            time_estimate_seconds = int(estimated_seconds)
            time_estimate_display = f"{time_estimate_minutes:.1f} minutes"

        return {
            "cell_count": cell_count,
            "estimated_cost_usd": round(estimated_cost, 4),
            "estimated_time_seconds": time_estimate_seconds,
            "estimated_time_display": time_estimate_display,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Estimation failed: {str(e)}")
    finally:
        input_path.unlink(missing_ok=True)


@app.post("/sheets")
async def get_sheets(
    file: UploadFile = File(..., description="Excel file to get sheet names from"),
) -> dict:
    """Get sheet names from an Excel file.

    Returns a list of sheet names in the uploaded Excel file.
    """
    # Validate file type
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")

    if not file.filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(
            status_code=400,
            detail="Invalid file type. Only Excel files (.xlsx, .xlsm, .xltx, .xltm) are supported",
        )

    # Read file content
    content = await file.read()

    # Check file size
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=400,
            detail=f"File too large. Maximum size is {MAX_FILE_SIZE // (1024 * 1024)}MB",
        )

    # Save to temp file for processing
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_input:
        tmp_input.write(content)
        input_path = Path(tmp_input.name)

    try:
        # Load workbook and get sheet names
        wb = load_workbook(input_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()

        return {"sheets": sheet_names}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read file: {str(e)}")
    finally:
        input_path.unlink(missing_ok=True)


def col_to_letter(col: int) -> str:
    """Convert column number to Excel letter (1 -> A, 27 -> AA)."""
    result = ""
    while col > 0:
        col, remainder = divmod(col - 1, 26)
        result = chr(65 + remainder) + result
    return result


@app.post("/count")
async def count_translatable(
    file: UploadFile = File(..., description="Excel file to count translatable cells"),
    sheets: Optional[str] = Form(None, description="Comma-separated sheet names (all if omitted)"),
) -> dict:
    """Count translatable cells in an Excel file.

    Returns the number of cells containing text that would be translated.
    Excludes formulas, numbers, dates, and empty cells.
    """
    # Validate file type
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")

    if not file.filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(
            status_code=400,
            detail="Invalid file type. Only Excel files (.xlsx, .xlsm, .xltx, .xltm) are supported",
        )

    # Read file content
    content = await file.read()

    # Check file size
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=400,
            detail=f"File too large. Maximum size is {MAX_FILE_SIZE // (1024 * 1024)}MB",
        )

    # Parse sheets parameter
    sheets_set = None
    if sheets:
        sheets_set = {s.strip() for s in sheets.split(",") if s.strip()}

    # Save to temp file for processing
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_input:
        tmp_input.write(content)
        input_path = Path(tmp_input.name)

    try:
        cell_count = count_cells(input_path, sheets_set)
        scope = f"sheets: {', '.join(sorted(sheets_set))}" if sheets_set else "all sheets"

        return {
            "cell_count": cell_count,
            "scope": scope,
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Count failed: {str(e)}")
    finally:
        input_path.unlink(missing_ok=True)


@app.post("/preview")
async def preview_cells(
    file: UploadFile = File(..., description="Excel file to preview"),
    limit: int = Form(10, ge=1, le=50, description="Maximum number of cells to preview (1-50)"),
    sheets: Optional[str] = Form(None, description="Comma-separated sheet names (all if omitted)"),
) -> dict:
    """Preview translatable cells from an Excel file.

    Returns the first N cells that would be translated, showing their
    location and content. Useful for understanding what will be translated
    before running a full translation.
    """
    # Validate file type
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")

    if not file.filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(
            status_code=400,
            detail="Invalid file type. Only Excel files (.xlsx, .xlsm, .xltx, .xltm) are supported",
        )

    # Read file content
    content = await file.read()

    # Check file size
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=400,
            detail=f"File too large. Maximum size is {MAX_FILE_SIZE // (1024 * 1024)}MB",
        )

    # Parse sheets parameter
    sheets_set = None
    if sheets:
        sheets_set = {s.strip() for s in sheets.split(",") if s.strip()}

    # Save to temp file for processing
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_input:
        tmp_input.write(content)
        input_path = Path(tmp_input.name)

    try:
        with ExcelExtractor(input_path, sheets=sheets_set) as extractor:
            cells = []
            for i, cell in enumerate(extractor.extract_cells()):
                if i >= limit:
                    break
                col_letter = col_to_letter(cell.col)
                cells.append({
                    "sheet": cell.sheet,
                    "cell": f"{col_letter}{cell.row}",
                    "content": cell.value[:200] if len(cell.value) > 200 else cell.value,
                })

        scope = f"sheets: {', '.join(sorted(sheets_set))}" if sheets_set else "all sheets"

        return {
            "cells": cells,
            "total_shown": len(cells),
            "scope": scope,
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Preview failed: {str(e)}")
    finally:
        input_path.unlink(missing_ok=True)


@app.post("/translate")
async def translate(
    file: UploadFile = File(..., description="Excel file to translate"),
    target_lang: str = Form(..., description="Target language (e.g., french, spanish)"),
    source_lang: Optional[str] = Form(None, description="Source language (auto-detect if omitted)"),
    context: Optional[str] = Form(None, description="Additional context for accurate translations"),
    sheets: Optional[str] = Form(None, description="Comma-separated sheet names (all if omitted)"),
) -> FileResponse:
    """Translate an Excel file.

    Upload an Excel file and receive the translated version.
    Preserves all formatting, formulas, images, and data validations.
    """
    # Validate file type
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")

    if not file.filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(
            status_code=400,
            detail="Invalid file type. Only Excel files (.xlsx, .xlsm, .xltx, .xltm) are supported",
        )

    # Read file content
    content = await file.read()

    # Check file size
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=400,
            detail=f"File too large. Maximum size is {MAX_FILE_SIZE // (1024 * 1024)}MB",
        )

    # Parse sheets parameter
    sheets_set = None
    if sheets:
        sheets_set = {s.strip() for s in sheets.split(",") if s.strip()}

    # Save to temp file for processing
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_input:
        tmp_input.write(content)
        input_path = Path(tmp_input.name)

    try:
        # Check cell count
        cell_count = count_cells(input_path, sheets_set)
        if cell_count > MAX_CELLS:
            raise HTTPException(
                status_code=400,
                detail=f"Too many cells ({cell_count}). Maximum is {MAX_CELLS} cells per request",
            )

        if cell_count == 0:
            raise HTTPException(
                status_code=400,
                detail="No translatable content found in the file",
            )

        # Create output path
        output_path = input_path.with_name(f"{input_path.stem}_translated.xlsx")

        # Translate (without progress callback for now - can be added with SSE)
        result = translate_file(
            input_file=input_path,
            output_file=output_path,
            target_lang=target_lang,
            source_lang=source_lang,
            context=context,
            sheets=sheets_set,
        )

        # Return translated file
        output_filename = file.filename.replace(".xlsx", f"_{target_lang}.xlsx")
        return FileResponse(
            path=output_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"X-Cells-Translated": str(result["cell_count"])},
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Translation failed: {str(e)}")
    finally:
        # Cleanup input file (output file cleaned up after response is sent)
        input_path.unlink(missing_ok=True)


@app.post("/translate-stream")
async def translate_stream(
    file: UploadFile = File(..., description="Excel file to translate"),
    target_lang: str = Form(..., description="Target language (e.g., french, spanish)"),
    source_lang: Optional[str] = Form(None, description="Source language (auto-detect if omitted)"),
    context: Optional[str] = Form(None, description="Additional context for accurate translations"),
    sheets: Optional[str] = Form(None, description="Comma-separated sheet names (all if omitted)"),
):
    """Translate an Excel file with real-time progress streaming via SSE.

    Returns Server-Sent Events with progress updates and the final translated file.
    """
    # Validate file type
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")

    if not file.filename.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        raise HTTPException(
            status_code=400,
            detail="Invalid file type. Only Excel files (.xlsx, .xlsm, .xltx, .xltm) are supported",
        )

    # Read file content
    content = await file.read()

    # Check file size
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=400,
            detail=f"File too large. Maximum size is {MAX_FILE_SIZE // (1024 * 1024)}MB",
        )

    # Parse sheets parameter
    sheets_set = None
    if sheets:
        sheets_set = {s.strip() for s in sheets.split(",") if s.strip()}

    # Save to temp file for processing
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_input:
        tmp_input.write(content)
        input_path = Path(tmp_input.name)

    # Check cell count
    try:
        cell_count = count_cells(input_path, sheets_set)
        if cell_count > MAX_CELLS:
            input_path.unlink(missing_ok=True)
            raise HTTPException(
                status_code=400,
                detail=f"Too many cells ({cell_count}). Maximum is {MAX_CELLS} cells per request",
            )

        if cell_count == 0:
            input_path.unlink(missing_ok=True)
            raise HTTPException(
                status_code=400,
                detail="No translatable content found in the file",
            )
    except HTTPException:
        raise
    except Exception as e:
        input_path.unlink(missing_ok=True)
        raise HTTPException(status_code=500, detail=f"File validation failed: {str(e)}")

    # Create output path
    output_path = input_path.with_name(f"{input_path.stem}_translated.xlsx")
    output_filename = file.filename.replace(".xlsx", f"_{target_lang}.xlsx")

    # Progress queue for thread communication
    progress_queue: queue.Queue = queue.Queue()
    result_holder = {"result": None, "error": None}

    def progress_callback(translated: int, total: int, stage: str):
        """Callback to report translation progress."""
        progress_queue.put({
            "translated": translated,
            "total": total,
            "stage": stage,
            "percentage": round((translated / total) * 100) if total > 0 else 0
        })

    def run_translation():
        """Run translation in a separate thread."""
        try:
            result_holder["result"] = translate_file(
                input_file=input_path,
                output_file=output_path,
                target_lang=target_lang,
                source_lang=source_lang,
                context=context,
                sheets=sheets_set,
                progress_callback=progress_callback,
            )
        except Exception as e:
            result_holder["error"] = str(e)

    # Start translation in background thread
    thread = threading.Thread(target=run_translation)
    thread.start()

    async def event_generator():
        """Generate SSE events for progress and completion."""
        try:
            while thread.is_alive() or not progress_queue.empty():
                try:
                    progress = progress_queue.get(timeout=0.1)
                    yield {
                        "event": "progress",
                        "data": json.dumps(progress)
                    }
                except queue.Empty:
                    await asyncio.sleep(0.1)

            # Wait for thread to complete
            thread.join()

            # Send final result or error
            if result_holder["error"]:
                yield {
                    "event": "error",
                    "data": json.dumps({"error": result_holder["error"]})
                }
            else:
                # Encode translated file as base64 for SSE
                with open(output_path, "rb") as f:
                    file_base64 = base64.b64encode(f.read()).decode()
                yield {
                    "event": "complete",
                    "data": json.dumps({
                        "filename": output_filename,
                        "file_base64": file_base64,
                        "cell_count": result_holder["result"]["cell_count"]
                    })
                }
        finally:
            # Cleanup temp files
            input_path.unlink(missing_ok=True)
            output_path.unlink(missing_ok=True)

    return EventSourceResponse(event_generator())
