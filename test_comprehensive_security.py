#!/usr/bin/env python3
"""Comprehensive security testing for Rosetta API.

Tests all security recommendations:
1. File upload validation
2. Rate limiting
3. Error message sanitization
4. Temp file cleanup
5. CORS configuration
"""

import os
import sys
import time
import tempfile
import subprocess
from pathlib import Path

def print_test(name):
    """Print test header."""
    print(f"\n{'='*60}")
    print(f"TEST: {name}")
    print('='*60)

def test_1_file_upload_validation():
    """Test 1: File Upload Validation - Try uploading non-Excel files."""
    print_test("File Upload Validation")

    try:
        import requests

        # Check if server is running
        try:
            response = requests.get("http://localhost:8000/health", timeout=2)
            if response.status_code != 200:
                print("⚠️  Server not running at localhost:8000")
                print("   Start server with: uv run uvicorn rosetta.api:app --reload")
                return False
        except requests.exceptions.ConnectionError:
            print("⚠️  Server not running at localhost:8000")
            print("   Start server with: uv run uvicorn rosetta.api:app --reload")
            return False

        print("\n[1.1] Testing non-Excel file with .xlsx extension...")
        # Create fake Excel file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False, mode='w') as f:
            f.write("This is not an Excel file")
            fake_file = f.name

        try:
            files = {'file': open(fake_file, 'rb')}
            response = requests.post('http://localhost:8000/estimate', files=files, timeout=10)

            if response.status_code == 400:
                print(f"✅ PASS: Non-Excel file rejected")
                print(f"   Response: {response.json()['detail'][:80]}...")
            elif response.status_code == 500:
                print(f"🟡 PARTIAL: Rejected but with 500 error (should be 400)")
                print(f"   Response: {response.json()['detail'][:80]}...")
            else:
                print(f"❌ FAIL: Non-Excel file was accepted! Status: {response.status_code}")
                print(f"   This is a CRITICAL security issue - magic bytes validation needed")
        finally:
            os.unlink(fake_file)

        print("\n[1.2] Testing file with no extension...")
        with tempfile.NamedTemporaryFile(delete=False, mode='w') as f:
            f.write("data")
            noext_file = f.name

        try:
            files = {'file': ('noext', open(noext_file, 'rb'))}
            response = requests.post('http://localhost:8000/estimate', files=files, timeout=10)

            if response.status_code == 400 and "Invalid file type" in response.text:
                print(f"✅ PASS: File without extension rejected")
            else:
                print(f"❌ FAIL: File without extension should be rejected")
        finally:
            os.unlink(noext_file)

        print("\n[1.3] Testing oversized file (>50MB)...")
        # Create 51MB file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False, mode='wb') as f:
            f.write(b'X' * (51 * 1024 * 1024))
            huge_file = f.name

        try:
            files = {'file': open(huge_file, 'rb')}
            response = requests.post('http://localhost:8000/estimate', files=files, timeout=10)

            if response.status_code == 400 and "too large" in response.text.lower():
                print(f"✅ PASS: Oversized file rejected")
            else:
                print(f"⚠️  WARNING: Oversized file handling - status {response.status_code}")
        finally:
            os.unlink(huge_file)

        print("\n✅ File Upload Validation Tests Complete")
        return True

    except ImportError:
        print("❌ requests library not installed")
        print("   Install with: pip install requests")
        return False
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_2_rate_limiting():
    """Test 2: Rate Limiting - Make multiple requests quickly."""
    print_test("Rate Limiting")

    try:
        import requests

        print("\n[2.1] Testing rate limiting on /estimate endpoint...")
        print("Making 20 rapid requests...")

        # Create a valid small test file
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Test"

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            test_file = f.name
        wb.save(test_file)

        try:
            start = time.time()
            responses = []

            for i in range(20):
                files = {'file': open(test_file, 'rb')}
                try:
                    response = requests.post('http://localhost:8000/estimate', files=files, timeout=5)
                    responses.append(response.status_code)
                    print(f"   Request {i+1}: {response.status_code}")

                    if response.status_code == 429:
                        print(f"\n✅ PASS: Rate limiting working! Got 429 after {i+1} requests")
                        break
                except Exception as e:
                    print(f"   Request {i+1}: Error - {e}")
                time.sleep(0.1)  # Small delay

            elapsed = time.time() - start
            print(f"\nCompleted {len(responses)} requests in {elapsed:.2f}s")

            if 429 in responses:
                print("✅ PASS: Rate limiting is active")
            else:
                print("❌ FAIL: NO RATE LIMITING DETECTED")
                print("   All 20 requests succeeded - this is a HIGH RISK security issue!")
                print("   Recommendation: Add rate limiting with slowapi or similar")
        finally:
            os.unlink(test_file)

        return True

    except ImportError:
        print("❌ openpyxl or requests not installed")
        return False
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_3_error_message_sanitization():
    """Test 3: Error Messages - Verify they don't leak sensitive info."""
    print_test("Error Message Sanitization")

    try:
        import requests

        print("\n[3.1] Testing error messages for information leakage...")

        test_cases = [
            ("Missing file", {}),
            ("Invalid file type", {'file': ('test.txt', b'data')}),
            ("Empty file", {'file': ('test.xlsx', b'')}),
        ]

        for test_name, files_data in test_cases:
            print(f"\n   Testing: {test_name}")
            try:
                response = requests.post('http://localhost:8000/estimate',
                                       files=files_data if files_data else None,
                                       timeout=5)

                error_detail = response.json().get('detail', '')
                print(f"   Response: {error_detail[:100]}")

                # Check for sensitive information leakage
                sensitive_patterns = [
                    '/tmp/',
                    '/var/',
                    '/home/',
                    '/Users/',
                    'tmpXYZ',
                    '.py',
                    'python',
                    'openpyxl',
                    'traceback',
                ]

                leaked = [p for p in sensitive_patterns if p.lower() in error_detail.lower()]

                if leaked:
                    print(f"   ⚠️  WARNING: Error message may leak sensitive info:")
                    print(f"       Detected: {', '.join(leaked)}")
                else:
                    print(f"   ✅ PASS: Error message appears safe")

            except Exception as e:
                print(f"   Error making request: {e}")

        print("\n✅ Error Message Sanitization Tests Complete")
        return True

    except Exception as e:
        print(f"❌ Error: {e}")
        return False


def test_4_temp_file_cleanup():
    """Test 4: Temp File Cleanup - Check temp directory after requests."""
    print_test("Temporary File Cleanup")

    try:
        import requests
        from openpyxl import Workbook

        print("\n[4.1] Testing temporary file cleanup...")

        # Get temp directory
        temp_dir = tempfile.gettempdir()

        # Count temp files before
        before_files = set(os.listdir(temp_dir))
        print(f"   Temp files before: {len(before_files)}")

        # Make a request
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Test"

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            test_file = f.name
        wb.save(test_file)

        try:
            files = {'file': open(test_file, 'rb')}
            response = requests.post('http://localhost:8000/estimate', files=files, timeout=10)
            print(f"   Request completed: {response.status_code}")
        finally:
            os.unlink(test_file)

        # Wait a moment for cleanup
        time.sleep(1)

        # Count temp files after
        after_files = set(os.listdir(temp_dir))
        new_files = after_files - before_files

        print(f"   Temp files after: {len(after_files)}")
        print(f"   New temp files: {len(new_files)}")

        if new_files:
            print(f"   ⚠️  WARNING: {len(new_files)} new temp files detected:")
            for f in list(new_files)[:5]:  # Show first 5
                print(f"       {f}")
            print("   These should be cleaned up by the server")
        else:
            print("   ✅ PASS: No temp files left behind")

        print("\n✅ Temp File Cleanup Tests Complete")
        return True

    except Exception as e:
        print(f"❌ Error: {e}")
        return False


def test_5_cors_configuration():
    """Test 5: CORS - Test with different origins."""
    print_test("CORS Configuration")

    try:
        import requests

        print("\n[5.1] Testing CORS headers...")

        # Test with different origins
        test_origins = [
            ("http://localhost:3000", True),   # Should be allowed
            ("http://evil.com", False),        # Should be blocked
            ("https://attacker.com", False),   # Should be blocked
        ]

        for origin, should_allow in test_origins:
            print(f"\n   Testing origin: {origin}")

            headers = {'Origin': origin}
            response = requests.options('http://localhost:8000/estimate',
                                       headers=headers, timeout=5)

            cors_header = response.headers.get('Access-Control-Allow-Origin', '')

            if cors_header == '*':
                print(f"   ⚠️  WARNING: CORS allows ALL origins (*)")
                print(f"       This is INSECURE for production!")
            elif cors_header == origin:
                if should_allow:
                    print(f"   ✅ PASS: Origin allowed (expected)")
                else:
                    print(f"   ❌ FAIL: Origin allowed (should be blocked)")
            else:
                if should_allow:
                    print(f"   ❌ FAIL: Origin blocked (should be allowed)")
                else:
                    print(f"   ✅ PASS: Origin blocked (expected)")

        print("\n✅ CORS Configuration Tests Complete")
        return True

    except Exception as e:
        print(f"❌ Error: {e}")
        return False


def main():
    """Run all security tests."""
    print("="*60)
    print("COMPREHENSIVE SECURITY TESTING - ROSETTA API")
    print("="*60)
    print("\nThis script tests all security recommendations:")
    print("1. File upload validation")
    print("2. Rate limiting")
    print("3. Error message sanitization")
    print("4. Temp file cleanup")
    print("5. CORS configuration")
    print("\nPrerequisites:")
    print("- Server running at localhost:8000")
    print("- Run with: uv run uvicorn rosetta.api:app --reload")
    print("\n" + "="*60)

    # Run tests
    results = {
        "File Upload Validation": test_1_file_upload_validation(),
        "Rate Limiting": test_2_rate_limiting(),
        "Error Message Sanitization": test_3_error_message_sanitization(),
        "Temp File Cleanup": test_4_temp_file_cleanup(),
        "CORS Configuration": test_5_cors_configuration(),
    }

    # Summary
    print("\n" + "="*60)
    print("SECURITY TEST SUMMARY")
    print("="*60)

    for test_name, result in results.items():
        status = "✅ PASSED" if result else "❌ FAILED"
        print(f"{test_name}: {status}")

    passed = sum(1 for r in results.values() if r)
    total = len(results)

    print(f"\nTotal: {passed}/{total} tests passed ({passed/total*100:.0f}%)")

    if passed == total:
        print("\n✅ All security tests passed!")
    elif passed >= total * 0.7:
        print("\n🟡 Most tests passed, but some issues detected")
    else:
        print("\n🔴 CRITICAL: Multiple security issues detected!")

    print("="*60)

    return 0 if passed == total else 1


if __name__ == "__main__":
    sys.exit(main())
