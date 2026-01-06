"""Tests for MCP HTTP endpoints."""

import base64
import io
from unittest.mock import patch, MagicMock

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from rosetta.api import app


@pytest.fixture
def client():
    """Create a test client for the MCP API."""
    return TestClient(app)


@pytest.fixture
def sample_excel_base64():
    """Create a simple Excel file and return as base64."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Hello"
    ws["A2"] = "World"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    content = buffer.getvalue()
    return base64.b64encode(content).decode()


@pytest.fixture
def large_excel_base64():
    """Create a large Excel file (>50MB) and return as base64."""
    # Create content that exceeds 50MB when base64 encoded
    large_content = b"x" * (51 * 1024 * 1024)
    return base64.b64encode(large_content).decode()


class TestMCPInfoEndpoint:
    """Tests for GET /mcp/ endpoint."""

    def test_info_endpoint_returns_server_info(self, client):
        """GET /mcp/ should return server information."""
        response = client.get("/mcp/")
        assert response.status_code == 200
        data = response.json()
        assert "name" in data
        assert "version" in data
        assert "description" in data
        assert "endpoints" in data
        assert "tools" in data
        assert data["name"] == "Rosetta MCP Server"
        assert isinstance(data["tools"], list)
        assert len(data["tools"]) > 0


class TestMCPInitializeEndpoint:
    """Tests for POST /mcp/initialize endpoint."""

    def test_initialize_returns_protocol_info(self, client):
        """POST /mcp/initialize should return protocol version and capabilities."""
        response = client.post("/mcp/initialize")
        assert response.status_code == 200
        data = response.json()
        assert "protocolVersion" in data
        assert "capabilities" in data
        assert "serverInfo" in data
        assert data["protocolVersion"] == "2024-11-05"
        assert data["serverInfo"]["name"] == "rosetta"
        assert data["serverInfo"]["version"] == "0.1.0"


class TestMCPToolsListEndpoint:
    """Tests for GET /mcp/tools endpoint."""

    def test_list_tools_returns_all_tools(self, client):
        """GET /mcp/tools should return list of available tools."""
        response = client.get("/mcp/tools")
        assert response.status_code == 200
        data = response.json()
        assert "tools" in data
        assert isinstance(data["tools"], list)
        assert len(data["tools"]) > 0

        # Check tool structure
        tool = data["tools"][0]
        assert "name" in tool
        assert "description" in tool
        assert "inputSchema" in tool
        assert "type" in tool["inputSchema"]
        assert "properties" in tool["inputSchema"]

    def test_tools_include_expected_names(self, client):
        """Tools list should include expected tool names."""
        response = client.get("/mcp/tools")
        data = response.json()
        tool_names = [tool["name"] for tool in data["tools"]]
        assert "translate_excel" in tool_names
        assert "get_excel_sheets" in tool_names
        assert "count_translatable_cells" in tool_names
        assert "preview_cells" in tool_names
        assert "estimate_translation_cost" in tool_names


class TestMCPToolCallEndpoint:
    """Tests for POST /mcp/tools/call endpoint."""

    def test_call_unknown_tool_returns_error(self, client):
        """Calling unknown tool should return 400 error."""
        response = client.post(
            "/mcp/tools/call",
            json={
                "name": "unknown_tool",
                "arguments": {},
            },
        )
        assert response.status_code == 400
        assert "Unknown tool" in response.json()["detail"]

    def test_call_tool_without_name_returns_422(self, client):
        """Calling without tool name should return 422."""
        response = client.post(
            "/mcp/tools/call",
            json={
                "arguments": {},
            },
        )
        assert response.status_code == 422

    def test_call_tool_with_invalid_arguments_returns_error(self, client, sample_excel_base64):
        """Calling tool with invalid arguments should return error."""
        response = client.post(
            "/mcp/tools/call",
            json={
                "name": "translate_excel",
                "arguments": {
                    "file_content_base64": "invalid-base64!!!",
                    "filename": "test.xlsx",
                    "target_language": "french",
                },
            },
        )
        assert response.status_code == 200  # MCP returns 200 with isError flag
        data = response.json()
        assert data["isError"] is True
        assert "Validation" in data["content"][0]["text"] or "error" in data["content"][0]["text"].lower()

    def test_call_get_sheets_with_valid_file(self, client, sample_excel_base64):
        """Calling get_excel_sheets with valid file should return sheet names."""
        response = client.post(
            "/mcp/tools/call",
            json={
                "name": "get_excel_sheets",
                "arguments": {
                    "file_content_base64": sample_excel_base64,
                },
            },
        )
        assert response.status_code == 200
        data = response.json()
        assert data["isError"] is False
        assert "content" in data
        assert len(data["content"]) > 0
        assert "sheet" in data["content"][0]["text"].lower()

    def test_call_count_cells_with_valid_file(self, client, sample_excel_base64):
        """Calling count_translatable_cells with valid file should return count."""
        response = client.post(
            "/mcp/tools/call",
            json={
                "name": "count_translatable_cells",
                "arguments": {
                    "file_content_base64": sample_excel_base64,
                },
            },
        )
        assert response.status_code == 200
        data = response.json()
        assert data["isError"] is False
        assert "content" in data
        assert "cells" in data["content"][0]["text"].lower()

    def test_call_preview_cells_with_valid_file(self, client, sample_excel_base64):
        """Calling preview_cells with valid file should return preview."""
        response = client.post(
            "/mcp/tools/call",
            json={
                "name": "preview_cells",
                "arguments": {
                    "file_content_base64": sample_excel_base64,
                    "limit": 5,
                },
            },
        )
        assert response.status_code == 200
        data = response.json()
        assert data["isError"] is False
        assert "content" in data
        assert "preview" in data["content"][0]["text"].lower() or "cell" in data["content"][0]["text"].lower()

    def test_call_estimate_cost_with_valid_file(self, client, sample_excel_base64):
        """Calling estimate_translation_cost with valid file should return estimate."""
        response = client.post(
            "/mcp/tools/call",
            json={
                "name": "estimate_translation_cost",
                "arguments": {
                    "file_content_base64": sample_excel_base64,
                },
            },
        )
        assert response.status_code == 200
        data = response.json()
        assert data["isError"] is False
        assert "content" in data
        assert "cost" in data["content"][0]["text"].lower() or "estimate" in data["content"][0]["text"].lower()

    def test_call_translate_excel_missing_required_field(self, client, sample_excel_base64):
        """Calling translate_excel without required fields should return error."""
        response = client.post(
            "/mcp/tools/call",
            json={
                "name": "translate_excel",
                "arguments": {
                    "file_content_base64": sample_excel_base64,
                    # Missing filename and target_language
                },
            },
        )
        assert response.status_code == 200
        data = response.json()
        assert data["isError"] is True
        assert "Validation" in data["content"][0]["text"] or "required" in data["content"][0]["text"].lower()

    def test_call_tool_with_empty_arguments(self, client):
        """Calling tool with empty arguments should handle gracefully."""
        response = client.post(
            "/mcp/tools/call",
            json={
                "name": "get_excel_sheets",
                "arguments": {},
            },
        )
        assert response.status_code == 200
        data = response.json()
        # Should return validation error
        assert data["isError"] is True

    def test_call_tool_with_malformed_json(self, client):
        """Calling with malformed JSON should return 422."""
        response = client.post(
            "/mcp/tools/call",
            data="not json",
            headers={"Content-Type": "application/json"},
        )
        assert response.status_code == 422

    def test_error_response_format(self, client):
        """Error responses should follow MCP format with isError flag."""
        response = client.post(
            "/mcp/tools/call",
            json={
                "name": "get_excel_sheets",
                "arguments": {
                    "file_content_base64": "invalid",
                },
            },
        )
        assert response.status_code == 200
        data = response.json()
        assert "isError" in data
        assert "content" in data
        assert isinstance(data["content"], list)
        assert len(data["content"]) > 0
        assert "type" in data["content"][0]
        assert "text" in data["content"][0]

