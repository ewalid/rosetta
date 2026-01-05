"""Tests for MCP tool implementations."""

import base64
import io
from pathlib import Path
from unittest.mock import patch, MagicMock, mock_open

import pytest
from openpyxl import Workbook

from rosetta.api.mcp import (
    tool_translate_excel,
    tool_get_sheets,
    tool_count_cells,
    tool_preview_cells,
    tool_estimate_cost,
    MCPToolCallResult,
    MCPContentItem,
)
from rosetta.models import Cell


@pytest.fixture
def sample_excel_base64():
    """Create a simple Excel file and return as base64."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Hello"
    ws["A2"] = "World"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    content = buffer.getvalue()
    return base64.b64encode(content).decode()


@pytest.fixture
def empty_excel_base64():
    """Create an Excel file with no text content."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = 123  # Number
    ws["A2"] = "=SUM(1,2)"  # Formula

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    content = buffer.getvalue()
    return base64.b64encode(content).decode()


@pytest.fixture
def multi_sheet_excel_base64():
    """Create an Excel file with multiple sheets."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = "Hello"

    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "Bonjour"

    ws3 = wb.create_sheet("Sheet3")
    ws3["A1"] = "Hola"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    content = buffer.getvalue()
    return base64.b64encode(content).decode()


class TestToolTranslateExcel:
    """Tests for tool_translate_excel function."""

    @patch("rosetta.api.mcp.translate_file")
    @patch("rosetta.api.mcp.count_cells")
    @patch("rosetta.api.mcp.decode_file_to_temp")
    @patch("rosetta.api.mcp.encode_file_to_base64")
    def test_successful_translation(
        self, mock_encode, mock_decode, mock_count, mock_translate, sample_excel_base64, tmp_path
    ):
        """Test successful translation returns correct result."""
        # Setup mocks
        input_path = tmp_path / "input.xlsx"
        output_path = tmp_path / "input_french.xlsx"
        mock_decode.return_value = input_path
        mock_count.return_value = 2
        mock_translate.return_value = {
            "cell_count": 2,
            "rich_text_cells": 0,
            "dropdown_count": 0,
        }
        mock_encode.return_value = "encoded_output_base64"

        args = {
            "file_content_base64": sample_excel_base64,
            "filename": "test.xlsx",
            "target_language": "french",
        }

        result = tool_translate_excel(args)

        assert isinstance(result, MCPToolCallResult)
        assert result.isError is False
        assert len(result.content) > 0
        assert "Translation complete" in result.content[0].text
        assert "french" in result.content[0].text
        mock_translate.assert_called_once()

    @patch("rosetta.api.mcp.count_cells")
    @patch("rosetta.api.mcp.decode_file_to_temp")
    def test_empty_file_returns_error(self, mock_decode, mock_count, empty_excel_base64, tmp_path):
        """Test that empty file returns error."""
        input_path = tmp_path / "input.xlsx"
        mock_decode.return_value = input_path
        mock_count.return_value = 0

        args = {
            "file_content_base64": empty_excel_base64,
            "filename": "test.xlsx",
            "target_language": "french",
        }

        result = tool_translate_excel(args)

        assert result.isError is True
        assert "No translatable content" in result.content[0].text

    @patch("rosetta.api.mcp.count_cells")
    @patch("rosetta.api.mcp.decode_file_to_temp")
    def test_too_many_cells_returns_error(self, mock_decode, mock_count, sample_excel_base64, tmp_path):
        """Test that file with too many cells returns error."""
        input_path = tmp_path / "input.xlsx"
        mock_decode.return_value = input_path
        mock_count.return_value = 5001  # Exceeds limit

        args = {
            "file_content_base64": sample_excel_base64,
            "filename": "test.xlsx",
            "target_language": "french",
        }

        result = tool_translate_excel(args)

        assert result.isError is True
        assert "exceeds the limit" in result.content[0].text or "5000" in result.content[0].text

    @patch("rosetta.api.mcp.translate_file")
    @patch("rosetta.api.mcp.count_cells")
    @patch("rosetta.api.mcp.decode_file_to_temp")
    @patch("rosetta.api.mcp.encode_file_to_base64")
    def test_translation_with_optional_params(
        self, mock_encode, mock_decode, mock_count, mock_translate, sample_excel_base64, tmp_path
    ):
        """Test translation with optional parameters."""
        input_path = tmp_path / "input.xlsx"
        output_path = tmp_path / "input_french.xlsx"
        mock_decode.return_value = input_path
        mock_count.return_value = 2
        mock_translate.return_value = {
            "cell_count": 2,
            "rich_text_cells": 1,
            "dropdown_count": 1,
        }
        mock_encode.return_value = "encoded_output_base64"

        args = {
            "file_content_base64": sample_excel_base64,
            "filename": "test.xlsx",
            "target_language": "spanish",
            "source_language": "english",
            "context": "Medical terminology",
            "sheets": ["Sheet1"],
        }

        result = tool_translate_excel(args)

        assert result.isError is False
        # Verify translate_file was called with correct parameters
        call_kwargs = mock_translate.call_args.kwargs
        assert call_kwargs["target_lang"] == "spanish"
        assert call_kwargs["source_lang"] == "english"
        assert call_kwargs["context"] == "Medical terminology"
        assert call_kwargs["sheets"] == {"Sheet1"}


class TestToolGetSheets:
    """Tests for tool_get_sheets function."""

    @patch("rosetta.api.mcp.ExcelExtractor")
    @patch("rosetta.api.mcp.decode_file_to_temp")
    def test_get_single_sheet(self, mock_decode, mock_extractor_class, sample_excel_base64, tmp_path):
        """Test getting sheets from file with single sheet."""
        input_path = tmp_path / "input.xlsx"
        mock_decode.return_value = input_path

        mock_extractor = MagicMock()
        mock_extractor.sheet_names = ["Sheet1"]
        mock_extractor_class.return_value.__enter__.return_value = mock_extractor

        args = {
            "file_content_base64": sample_excel_base64,
        }

        result = tool_get_sheets(args)

        assert isinstance(result, MCPToolCallResult)
        assert result.isError is False
        assert "Sheet1" in result.content[0].text

    @patch("rosetta.api.mcp.ExcelExtractor")
    @patch("rosetta.api.mcp.decode_file_to_temp")
    def test_get_multiple_sheets(self, mock_decode, mock_extractor_class, multi_sheet_excel_base64, tmp_path):
        """Test getting sheets from file with multiple sheets."""
        input_path = tmp_path / "input.xlsx"
        mock_decode.return_value = input_path

        mock_extractor = MagicMock()
        mock_extractor.sheet_names = ["Sheet1", "Sheet2", "Sheet3"]
        mock_extractor_class.return_value.__enter__.return_value = mock_extractor

        args = {
            "file_content_base64": multi_sheet_excel_base64,
        }

        result = tool_get_sheets(args)

        assert result.isError is False
        assert "Sheet1" in result.content[0].text
        assert "Sheet2" in result.content[0].text
        assert "Sheet3" in result.content[0].text


class TestToolCountCells:
    """Tests for tool_count_cells function."""

    @patch("rosetta.api.mcp.count_cells")
    @patch("rosetta.api.mcp.decode_file_to_temp")
    def test_count_all_sheets(self, mock_decode, mock_count, sample_excel_base64, tmp_path):
        """Test counting cells from all sheets."""
        input_path = tmp_path / "input.xlsx"
        mock_decode.return_value = input_path
        mock_count.return_value = 5

        args = {
            "file_content_base64": sample_excel_base64,
        }

        result = tool_count_cells(args)

        assert result.isError is False
        assert "5" in result.content[0].text
        mock_count.assert_called_once_with(input_path, None)

    @patch("rosetta.api.mcp.count_cells")
    @patch("rosetta.api.mcp.decode_file_to_temp")
    def test_count_specific_sheets(self, mock_decode, mock_count, sample_excel_base64, tmp_path):
        """Test counting cells from specific sheets."""
        input_path = tmp_path / "input.xlsx"
        mock_decode.return_value = input_path
        mock_count.return_value = 3

        args = {
            "file_content_base64": sample_excel_base64,
            "sheets": ["Sheet1", "Sheet2"],
        }

        result = tool_count_cells(args)

        assert result.isError is False
        assert "3" in result.content[0].text
        # Verify sheets parameter was converted to set
        # count_cells is called with (input_path, sheets)
        call_args = mock_count.call_args[0]  # positional arguments
        assert call_args[1] == {"Sheet1", "Sheet2"}


class TestToolPreviewCells:
    """Tests for tool_preview_cells function."""

    @patch("rosetta.api.mcp.ExcelExtractor")
    @patch("rosetta.api.mcp.decode_file_to_temp")
    def test_preview_with_limit(self, mock_decode, mock_extractor_class, sample_excel_base64, tmp_path):
        """Test previewing cells with limit."""
        input_path = tmp_path / "input.xlsx"
        mock_decode.return_value = input_path

        # Create mock cells
        mock_cells = [
            Cell(sheet="Sheet1", row=1, col=1, value="Hello"),
            Cell(sheet="Sheet1", row=2, col=1, value="World"),
        ]

        mock_extractor = MagicMock()
        mock_extractor.extract_cells.return_value = iter(mock_cells)
        mock_extractor_class.return_value.__enter__.return_value = mock_extractor

        args = {
            "file_content_base64": sample_excel_base64,
            "limit": 5,
        }

        result = tool_preview_cells(args)

        assert result.isError is False
        assert "preview" in result.content[0].text.lower() or "cell" in result.content[0].text.lower()
        assert "Hello" in result.content[0].text or "A1" in result.content[0].text

    @patch("rosetta.api.mcp.ExcelExtractor")
    @patch("rosetta.api.mcp.decode_file_to_temp")
    def test_preview_respects_limit(self, mock_decode, mock_extractor_class, sample_excel_base64, tmp_path):
        """Test that preview respects the limit parameter."""
        input_path = tmp_path / "input.xlsx"
        mock_decode.return_value = input_path

        # Create many mock cells
        mock_cells = [Cell(sheet="Sheet1", row=i, col=1, value=f"Cell{i}") for i in range(1, 21)]

        mock_extractor = MagicMock()
        mock_extractor.extract_cells.return_value = iter(mock_cells)
        mock_extractor_class.return_value.__enter__.return_value = mock_extractor

        args = {
            "file_content_base64": sample_excel_base64,
            "limit": 5,
        }

        result = tool_preview_cells(args)

        assert result.isError is False
        # Should only show 5 cells
        assert "5" in result.content[0].text

    @patch("rosetta.api.mcp.ExcelExtractor")
    @patch("rosetta.api.mcp.decode_file_to_temp")
    def test_preview_empty_file(self, mock_decode, mock_extractor_class, empty_excel_base64, tmp_path):
        """Test previewing empty file returns appropriate message."""
        input_path = tmp_path / "input.xlsx"
        mock_decode.return_value = input_path

        mock_extractor = MagicMock()
        mock_extractor.extract_cells.return_value = iter([])
        mock_extractor_class.return_value.__enter__.return_value = mock_extractor

        args = {
            "file_content_base64": empty_excel_base64,
            "limit": 10,
        }

        result = tool_preview_cells(args)

        assert result.isError is False
        assert "no translatable cells" in result.content[0].text.lower()

    @patch("rosetta.api.mcp.ExcelExtractor")
    @patch("rosetta.api.mcp.decode_file_to_temp")
    def test_preview_with_sheets_filter(self, mock_decode, mock_extractor_class, multi_sheet_excel_base64, tmp_path):
        """Test preview with sheet filtering."""
        input_path = tmp_path / "input.xlsx"
        mock_decode.return_value = input_path

        mock_cells = [
            Cell(sheet="Sheet1", row=1, col=1, value="Hello"),
        ]

        mock_extractor = MagicMock()
        mock_extractor.extract_cells.return_value = iter(mock_cells)
        mock_extractor_class.return_value.__enter__.return_value = mock_extractor

        args = {
            "file_content_base64": multi_sheet_excel_base64,
            "limit": 10,
            "sheets": ["Sheet1"],
        }

        result = tool_preview_cells(args)

        assert result.isError is False
        # Verify ExcelExtractor was called with sheets parameter
        call_kwargs = mock_extractor_class.call_args.kwargs
        assert call_kwargs["sheets"] == {"Sheet1"}


class TestToolEstimateCost:
    """Tests for tool_estimate_cost function."""

    @patch("rosetta.api.mcp.count_cells")
    @patch("rosetta.api.mcp.decode_file_to_temp")
    def test_estimate_cost_calculation(self, mock_decode, mock_count, sample_excel_base64, tmp_path):
        """Test cost estimation calculation."""
        input_path = tmp_path / "input.xlsx"
        mock_decode.return_value = input_path
        mock_count.return_value = 1000  # 1000 cells

        args = {
            "file_content_base64": sample_excel_base64,
        }

        result = tool_estimate_cost(args)

        assert result.isError is False
        assert "1000" in result.content[0].text
        assert "cost" in result.content[0].text.lower() or "estimate" in result.content[0].text.lower()

    @patch("rosetta.api.mcp.count_cells")
    @patch("rosetta.api.mcp.decode_file_to_temp")
    def test_estimate_cost_with_sheets(self, mock_decode, mock_count, sample_excel_base64, tmp_path):
        """Test cost estimation with specific sheets."""
        input_path = tmp_path / "input.xlsx"
        mock_decode.return_value = input_path
        mock_count.return_value = 500

        args = {
            "file_content_base64": sample_excel_base64,
            "sheets": ["Sheet1"],
        }

        result = tool_estimate_cost(args)

        assert result.isError is False
        assert "500" in result.content[0].text
        # Verify count_cells was called with sheets parameter
        call_args = mock_count.call_args[0]  # positional arguments
        assert call_args[1] == {"Sheet1"}

    @patch("rosetta.api.mcp.count_cells")
    @patch("rosetta.api.mcp.decode_file_to_temp")
    def test_estimate_cost_time_calculation(self, mock_decode, mock_count, sample_excel_base64, tmp_path):
        """Test that time estimate is included in response."""
        input_path = tmp_path / "input.xlsx"
        mock_decode.return_value = input_path
        mock_count.return_value = 2500  # Should take ~50 seconds

        args = {
            "file_content_base64": sample_excel_base64,
        }

        result = tool_estimate_cost(args)

        assert result.isError is False
        # Should include time estimate
        assert "time" in result.content[0].text.lower() or "minute" in result.content[0].text.lower() or "second" in result.content[0].text.lower()

